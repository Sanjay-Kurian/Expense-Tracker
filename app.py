from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
DATA_PATH = r'C:\Users\sanja\Documents\Projects\API\Final.xlsx'
REFERENCE_PATH = r'C:\Users\sanja\Documents\Projects\API\Categories.xlsx'

@app.route('/')
def index():
    df = pd.read_excel(DATA_PATH)
    missing = df[df['Category'].isna()]
    latest_date = df['Date'].max()
    with open('last_updated.txt', 'r') as f:
        lines = f.readlines()
        last_updated_date = lines[0].strip()

        if len(lines) > 1:
            last_updated_balance = float(lines[1].strip())
        else:
            last_updated_balance = float(input("Balance not found. Please enter the last known balance: "))
    df["Running Balance"] = df["Amount"].cumsum() + last_updated_balance
    latest_balance = df['Running Balance'].iloc[-1]
    with open('last_updated.txt', 'w') as f:
        f.write(latest_date.strftime('%Y-%m-%d') + '\n')
        f.write(str(latest_balance))
    return render_template('index.html', data=missing.to_dict(orient='records'), columns=missing.columns)

@app.route('/edit', methods=['GET', 'POST'])
def edit():
    df = pd.read_excel(DATA_PATH)

    if request.method == 'POST':
        for row in df[df['Category'].isna()].itertuples():
            new_cat = request.form.get(f'cat_{row.Index}')
            new_note = request.form.get(f'note_{row.Index}')
            if new_cat and new_note:
                df.at[row.Index, 'Category'] = new_cat
                df.at[row.Index, 'Note'] = new_note
        return redirect(url_for('index'))

    missing = df[df['Category'].isna()].copy()
    missing['Index'] = missing.index
    reference = pd.read_excel(REFERENCE_PATH).sort_values(by=['Category', 'Notes'])
    reference_data = reference.to_dict(orient='records')
    reference_columns = reference.columns.tolist()

    return render_template(
        'edit.html',
        data=missing.to_dict(orient='records'),
        columns=missing.columns.difference(['Index']),
        ref_data=reference_data,
        ref_columns=reference_columns
    )

@app.route('/add_note', methods=['POST'])
def add_note():
    note = request.form.get('new_note')
    category = request.form.get('new_category')

    if note and category:
        # Load reference file
        ref_df = pd.read_excel(REFERENCE_PATH)

        # Check for duplicates
        if not ((ref_df['Notes'] == note) & (ref_df['Category'] == category)).any():
            ref_df = pd.concat([ref_df, pd.DataFrame({'Notes': [note], 'Category': [category]})], ignore_index=True)
            ref_df.to_excel(REFERENCE_PATH, index=False)

    return redirect(url_for('edit'))

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/append')
def append_to_excel_table():
    cleaned_df = pd.read_excel(r"C:\Users\sanja\Documents\Projects\API\Final.xlsx")
    cleaned_df["Date"] = pd.to_datetime(cleaned_df["Date"]).dt.date
    filename = r'C:\Users\sanja\Documents\Projects\API\MM.xlsx'
    sheet_name = 'Sheet1' 
    table_name = 'MyTable'  

    wb = load_workbook(filename)
    ws = wb[sheet_name]

    existing_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_data.append(row)

    if not existing_data:
        return render_template('append.html', message = "No data in existing Excel file.")

    existing_df = pd.DataFrame(existing_data, columns=[cell.value for cell in ws[1]])

    # Create unique ID from key fields
    cleaned_df["unique_key"] = (
        cleaned_df["Date"].astype(str) + "|" +
        cleaned_df["Category"].astype(str) + "|" +
        cleaned_df["Note"].astype(str) + "|" +
        cleaned_df["Amount"].astype(str)
    )
    existing_df["unique_key"] = (
        existing_df["Date"].astype(str) + "|" +
        existing_df["Category"].astype(str) + "|" +
        existing_df["Note"].astype(str) + "|" +
        existing_df["Amount"].astype(str)
    )

    # Filter only new rows
    new_rows = cleaned_df[~cleaned_df["unique_key"].isin(existing_df["unique_key"])].copy()

    if new_rows.empty:
        return render_template('append.html', message="No new rows to append — everything is up to date.")

    new_rows.drop(columns=["unique_key"], inplace=True)
    cleaned_df.drop(columns="unique_key", inplace=True)

    table = ws.tables[table_name]
    table_range = table.ref 
    start_cell, end_cell = table_range.split(':')
    last_row = ws[end_cell].row

    for row in dataframe_to_rows(cleaned_df, index=False, header=False):
        ws.append(row)

    new_end_row = last_row + len(cleaned_df)
    table.ref = f"{start_cell.split(':')[0]}:{end_cell[:1]}{new_end_row}"

    wb.save(filename)
    return render_template('append.html', message=f"{len(new_rows)} new row(s) successfully appended.")

if __name__ == '__main__':
    app.run(debug=True)
